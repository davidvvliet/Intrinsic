import secrets
import io
import json
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from app.core.deps import get_workos_user
from app.core.limits import enforce_workspace_limit
from app.storage.async_db import execute_query, execute_query_one, execute_command
from app.api.templates import parse_xlsx_all_sheets, parse_csv, extract_preview_data, MAX_TEMPLATE_SIZE_BYTES
from pydantic import BaseModel
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, ScatterChart, AreaChart, Reference

router = APIRouter()

CHART_TYPE_MAP = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "doughnut": DoughnutChart,
    "scatter": ScatterChart,
    "area": AreaChart,
}

# Approximate pixel-to-cell conversion (default cell dimensions)
DEFAULT_COL_WIDTH_PX = 100
DEFAULT_ROW_HEIGHT_PX = 25


def _add_charts_to_sheet(ws, charts):
    print(f"[export] Adding {len(charts)} chart(s) to sheet '{ws.title}'")
    for chart_data in charts:
        chart_cls = CHART_TYPE_MAP.get(chart_data.get("type", "bar"), BarChart)
        chart = chart_cls()

        if chart_data.get("title"):
            chart.title = chart_data["title"]
        if chart_data.get("xAxisLabel") and hasattr(chart, "x_axis"):
            chart.x_axis.title = chart_data["xAxisLabel"]
        if chart_data.get("yAxisLabel") and hasattr(chart, "y_axis"):
            chart.y_axis.title = chart_data["yAxisLabel"]

        data_ranges = chart_data.get("dataRanges", [])
        if not data_ranges and "dataRange" in chart_data:
            dr = chart_data["dataRange"]
            data_ranges = [{
                "start": {"row": dr["startRow"], "col": dr["startCol"]},
                "end": {"row": dr["endRow"], "col": dr["endCol"]},
            }]
        if not data_ranges:
            print(f"[export] Chart '{chart_data.get('title', 'untitled')}' has no dataRanges, skipping")
            continue

        print(f"[export] Chart '{chart_data.get('title', 'untitled')}' has {len(data_ranges)} dataRange(s): {data_ranges}")

        use_headers = chart_data.get("useFirstRowAsHeaders", True)

        # First range = categories (x-axis labels)
        first = data_ranges[0]
        print(f"[export] Categories: row {first['start']['row']+1}-{first['end']['row']+1}, col {first['start']['col']+1}-{first['end']['col']+1}")
        cat_ref = Reference(
            ws,
            min_col=first["start"]["col"] + 1,
            min_row=first["start"]["row"] + 1,
            max_col=first["end"]["col"] + 1,
            max_row=first["end"]["row"] + 1,
        )

        # Remaining ranges = data series
        for dr in data_ranges[1:]:
            data_ref = Reference(
                ws,
                min_col=dr["start"]["col"] + 1,
                min_row=dr["start"]["row"] + 1,
                max_col=dr["end"]["col"] + 1,
                max_row=dr["end"]["row"] + 1,
            )
            print(f"[export] Data series: row {dr['start']['row']+1}-{dr['end']['row']+1}, col {dr['start']['col']+1}-{dr['end']['col']+1}")
            chart.add_data(data_ref, from_rows=True, titles_from_data=False)

        chart.set_categories(cat_ref)

        # Position: convert pixel coordinates to cell anchor
        pos = chart_data.get("position", {})
        anchor_col = max(1, int(pos.get("x", 0) / DEFAULT_COL_WIDTH_PX) + 1)
        anchor_row = max(1, int(pos.get("y", 0) / DEFAULT_ROW_HEIGHT_PX) + 1)
        anchor = f"{get_column_letter(anchor_col)}{anchor_row}"

        # Size
        if pos.get("width"):
            chart.width = pos["width"] / 7  # approximate px to cm
        if pos.get("height"):
            chart.height = pos["height"] / 7

        print(f"[export] Chart '{chart_data.get('title', 'untitled')}' type={chart_data.get('type')} ranges={len(data_ranges)} anchor={anchor}")
        ws.add_chart(chart, anchor)


class CreateWorkspaceRequest(BaseModel):
    id: Optional[str] = None
    name: Optional[str] = "Untitled"


class RenameWorkspaceRequest(BaseModel):
    name: str


@router.get("/workspaces")
async def list_workspaces(user=Depends(get_workos_user)):
    """List all workspaces for the user."""
    user_id = user["id"]

    rows = await execute_query(
        """SELECT id, name, thumbnail_url, preview_data, created_at, updated_at
           FROM workspaces
           WHERE user_id = $1
           ORDER BY updated_at DESC""",
        user_id
    )

    return [
        {
            "id": row["id"],
            "name": row["name"],
            "thumbnail_url": row["thumbnail_url"],
            "preview_data": row["preview_data"],
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
        }
        for row in rows
    ]


@router.post("/workspaces")
async def create_workspace(
    body: CreateWorkspaceRequest,
    user=Depends(get_workos_user)
):
    """Create a new workspace."""
    user_id = user["id"]
    await enforce_workspace_limit(user_id, user.get("email"))
    workspace_id = body.id or secrets.token_urlsafe(12)

    row = await execute_query_one(
        """INSERT INTO workspaces (id, user_id, name)
           VALUES ($1, $2, $3)
           RETURNING id, name, created_at, updated_at""",
        workspace_id, user_id, body.name or "Untitled"
    )

    return {
        "id": row["id"],
        "name": row["name"],
        "thumbnail_url": None,
        "created_at": row["created_at"].isoformat() if row["created_at"] else None,
        "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
    }


@router.get("/workspaces/{workspace_id}")
async def get_workspace(
    workspace_id: str,
    user=Depends(get_workos_user)
):
    """Get a workspace by ID."""
    user_id = user["id"]

    row = await execute_query_one(
        """SELECT id, name, thumbnail_url, created_at, updated_at
           FROM workspaces
           WHERE id = $1 AND user_id = $2""",
        workspace_id, user_id
    )

    if not row:
        raise HTTPException(status_code=404, detail="Workspace not found")

    return {
        "id": row["id"],
        "name": row["name"],
        "thumbnail_url": row["thumbnail_url"],
        "created_at": row["created_at"].isoformat() if row["created_at"] else None,
        "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
    }


@router.patch("/workspaces/{workspace_id}/name")
async def rename_workspace(
    workspace_id: str,
    body: RenameWorkspaceRequest,
    user=Depends(get_workos_user)
):
    """Rename a workspace."""
    user_id = user["id"]

    existing = await execute_query_one(
        "SELECT id FROM workspaces WHERE id = $1 AND user_id = $2",
        workspace_id, user_id
    )

    if not existing:
        raise HTTPException(status_code=404, detail="Workspace not found")

    await execute_command(
        "UPDATE workspaces SET name = $1, updated_at = NOW() WHERE id = $2 AND user_id = $3",
        body.name or "Untitled", workspace_id, user_id
    )

    return {"status": "renamed", "id": workspace_id}


@router.delete("/workspaces/{workspace_id}")
async def delete_workspace(
    workspace_id: str,
    user=Depends(get_workos_user)
):
    """Delete a workspace and all its sheets."""
    user_id = user["id"]

    existing = await execute_query_one(
        "SELECT id FROM workspaces WHERE id = $1 AND user_id = $2",
        workspace_id, user_id
    )

    if not existing:
        raise HTTPException(status_code=404, detail="Workspace not found")

    # Delete all sheets in the workspace first
    await execute_command(
        "DELETE FROM sheets WHERE workspace_id = $1 AND user_id = $2",
        workspace_id, user_id
    )

    # Delete the workspace
    await execute_command(
        "DELETE FROM workspaces WHERE id = $1 AND user_id = $2",
        workspace_id, user_id
    )

    return {"status": "deleted", "id": workspace_id}


@router.post("/workspaces/upload")
async def create_workspace_from_file(
    name: str = Form("Untitled"),
    file: UploadFile = File(...),
    user=Depends(get_workos_user)
):
    """Create a new workspace by uploading an xlsx or csv file."""
    user_id = user["id"]
    await enforce_workspace_limit(user_id, user.get("email"))

    if not file.filename:
        raise HTTPException(status_code=400, detail="Filename required")

    filename_lower = file.filename.lower()
    if not (filename_lower.endswith('.xlsx') or filename_lower.endswith('.csv')):
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are supported")

    content = await file.read()

    if len(content) > MAX_TEMPLATE_SIZE_BYTES * 10:
        raise HTTPException(status_code=413, detail="File too large. Max size is 10MB")

    try:
        if filename_lower.endswith('.xlsx'):
            sheets = parse_xlsx_all_sheets(content)
        else:
            sheets = parse_csv(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    workspace_name = name.strip() or file.filename.rsplit('.', 1)[0] or "Untitled"
    workspace_id = secrets.token_urlsafe(12)

    preview_data = None
    if sheets and sheets[0].get("data", {}).get("cells"):
        first_data = sheets[0]["data"]
        preview_data = extract_preview_data(first_data["cells"], first_data.get("formatting"))

    await execute_command(
        """INSERT INTO workspaces (id, user_id, name, preview_data)
           VALUES ($1, $2, $3, $4::jsonb)""",
        workspace_id, user_id, workspace_name, preview_data
    )

    created_sheets = []
    for sheet in sheets:
        sheet_id = secrets.token_urlsafe(12)
        await execute_command(
            """INSERT INTO sheets (id, workspace_id, user_id, name, data, updated_at)
               VALUES ($1, $2, $3, $4, $5::jsonb, NOW())""",
            sheet_id, workspace_id, user_id, sheet["name"], sheet["data"]
        )
        created_sheets.append({"id": sheet_id, "name": sheet["name"]})

    return {"workspace_id": workspace_id, "sheets": created_sheets}


@router.get("/workspaces/{workspace_id}/export")
async def export_workspace(
    workspace_id: str,
    user=Depends(get_workos_user)
):
    """Export workspace as XLSX file with all sheets."""
    user_id = user["id"]

    # Get workspace
    workspace = await execute_query_one(
        "SELECT id, name FROM workspaces WHERE id = $1 AND user_id = $2",
        workspace_id, user_id
    )

    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace not found")

    # Get all sheets with data
    sheets = await execute_query(
        """SELECT id, name, data
           FROM sheets
           WHERE workspace_id = $1 AND user_id = $2
           ORDER BY created_at ASC""",
        workspace_id, user_id
    )

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet in sheets:
        ws = wb.create_sheet(title=sheet["name"][:31])  # Excel sheet names max 31 chars

        data = sheet["data"]
        if isinstance(data, str):
            data = json.loads(data)

        cells = data.get("cells", {})

        for key, cell_data in cells.items():
            row_str, col_str = key.split(",")
            row = int(row_str) + 1  # Excel is 1-indexed
            col = int(col_str) + 1

            raw_value = cell_data.get("raw", "")
            cell_type = cell_data.get("type", "text")

            cell = ws.cell(row=row, column=col)

            if cell_type == "formula" and raw_value.startswith("="):
                cell.value = raw_value
            elif cell_type == "number":
                try:
                    cell.value = float(raw_value)
                except ValueError:
                    cell.value = raw_value
            else:
                cell.value = raw_value

        # Apply formatting
        formatting = data.get("formatting", {})
        for key, fmt in formatting.items():
            row_str, col_str = key.split(",")
            row = int(row_str) + 1
            col = int(col_str) + 1
            cell = ws.cell(row=row, column=col)

            # Font properties
            font_kwargs = {}
            if fmt.get("bold"):
                font_kwargs["bold"] = True
            if fmt.get("italic"):
                font_kwargs["italic"] = True
            if fmt.get("textColor"):
                font_kwargs["color"] = fmt["textColor"].lstrip("#")
            if font_kwargs:
                cell.font = Font(**font_kwargs)

            # Fill color
            if fmt.get("fillColor"):
                cell.fill = PatternFill(
                    start_color=fmt["fillColor"].lstrip("#"),
                    end_color=fmt["fillColor"].lstrip("#"),
                    fill_type="solid"
                )

            # Number format
            nf = fmt.get("numberFormat")
            if nf:
                nf_type = nf.get("type") if isinstance(nf, dict) else nf
                if nf_type == "currency":
                    cell.number_format = "$#,##0.00"
                elif nf_type == "percent":
                    cell.number_format = "0.00%"
                elif nf_type == "number":
                    cell.number_format = "#,##0.00"

        # Column widths
        settings = data.get("settings", {})
        for col_idx, width in settings.get("columnWidths", []):
            col_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[col_letter].width = width / 7  # px to approximate Excel units

        # Frozen panes
        frozen_rows = settings.get("frozenRows", 0)
        frozen_cols = settings.get("frozenColumns", 0)
        if frozen_rows or frozen_cols:
            ws.freeze_panes = ws.cell(row=frozen_rows + 1, column=frozen_cols + 1).coordinate

        # Charts
        sheet_charts = data.get("charts", [])
        if sheet_charts:
            _add_charts_to_sheet(ws, sheet_charts)

    # If no sheets, create empty one
    if len(wb.sheetnames) == 0:
        wb.create_sheet(title="Sheet 1")

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{workspace['name']}.xlsx"
    # Sanitize filename
    filename = "".join(c for c in filename if c.isalnum() or c in " ._-").strip() or "export.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
