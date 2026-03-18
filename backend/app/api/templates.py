import json
import re
import secrets
import io
import csv
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from app.core.deps import get_workos_user
from app.storage.async_db import execute_query, execute_query_one, execute_command
from pydantic import BaseModel
from typing import Optional, Dict, Any, List
from openpyxl import load_workbook

router = APIRouter()

MAX_TEMPLATE_SIZE_BYTES = 1 * 1024 * 1024  # 1MB


def extract_preview_data(cells: Dict[str, Any], formatting: Dict[str, Any] = None) -> Dict[str, Any]:
    """Extract A1:F10 (rows 0-9, cols 0-5) from cells for preview."""
    preview = {}
    for key, value in cells.items():
        parts = key.split(',')
        if len(parts) == 2:
            row, col = int(parts[0]), int(parts[1])
            if row < 10 and col < 6:
                entry = dict(value)
                if formatting and key in formatting:
                    entry['format'] = formatting[key]
                preview[key] = entry
    return preview if preview else None


def _parse_rgb(color) -> Optional[str]:
    """Convert an openpyxl Color object to a #RRGGBB hex string, or None if default/theme color."""
    try:
        if color is None or color.type == 'theme':
            return None
        rgb = color.rgb  # e.g. 'FF3F51B5' or '00000000'
        if not rgb or rgb in ('00000000', 'FF000000') and color.type == 'rgb':
            # Only skip pure black for text (it's the default); keep it for fill
            pass
        if len(rgb) == 8:
            return f"#{rgb[2:]}"  # strip alpha prefix
        if len(rgb) == 6:
            return f"#{rgb}"
    except Exception:
        pass
    return None


def _parse_number_format(fmt_code: str) -> Optional[Dict[str, Any]]:
    """Map an Excel number format code to our NumberFormatSettings."""
    if not fmt_code or fmt_code == 'General':
        return None
    f = fmt_code.lower()
    if '%' in f:
        return {"type": "percent"}
    if 'e+' in f or 'e-' in f:
        return {"type": "scientific"}
    if any(c in f for c in ('y', 'd')) and 'm' in f:
        if 'h' in f:
            return {"type": "datetime"}
        return {"type": "date"}
    if 'h' in f and ('m' in f or 's' in f):
        return {"type": "time"}
    if '$' in fmt_code or '£' in fmt_code or '€' in fmt_code or '[$' in fmt_code:
        # Try to extract currency symbol
        sym_match = re.search(r'\[(\$[^\]]*)\]', fmt_code)
        symbol = '$'
        if sym_match:
            symbol = sym_match.group(1).lstrip('$').split('-')[0] or '$'
            if not symbol:
                symbol = '$'
        # Check if rounded (no decimals)
        if '.' not in fmt_code:
            return {"type": "currencyRounded", "currencySymbol": symbol}
        return {"type": "currency", "currencySymbol": symbol}
    if '_(' in fmt_code or '_(* ' in fmt_code:
        return {"type": "accounting"}
    # Plain number — count decimal places
    dec_match = re.search(r'0\.(0+)', fmt_code)
    if dec_match:
        return {"type": "number", "decimals": len(dec_match.group(1))}
    if '0' in fmt_code or '#' in fmt_code:
        return {"type": "number", "decimals": 0}
    return None


def _parse_cell_format(cell) -> Optional[Dict[str, Any]]:
    """Extract CellFormat fields from an openpyxl cell."""
    fmt = {}
    try:
        font = cell.font
        if font:
            if font.bold:
                fmt['bold'] = True
            if font.italic:
                fmt['italic'] = True
            if font.strike:
                fmt['strikethrough'] = True
            if font.underline and font.underline != 'none':
                fmt['underline'] = True
            color = _parse_rgb(font.color)
            # Exclude default black text color
            if color and color.lower() not in ('#000000', '#ffffff' if False else ''):
                fmt['textColor'] = color
    except Exception:
        pass
    try:
        fill = cell.fill
        if fill and fill.fill_type == 'solid':
            color = _parse_rgb(fill.fgColor)
            if color and color.lower() != '#ffffff':
                fmt['fillColor'] = color
    except Exception:
        pass
    try:
        nf = _parse_number_format(cell.number_format)
        if nf:
            fmt['numberFormat'] = nf
    except Exception:
        pass
    return fmt if fmt else None


def parse_worksheet(worksheet) -> Dict[str, Any]:
    """Parse a single worksheet and convert to our cell format."""
    cells = {}
    formatting = {}
    max_row = 0
    max_col = 0

    for row in worksheet.iter_rows():
        for cell in row:
            row_idx = cell.row - 1
            col_idx = cell.column - 1

            max_row = max(max_row, row_idx)
            max_col = max(max_col, col_idx)

            key = f"{row_idx},{col_idx}"

            if cell.data_type == 'f' or (hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('=')):
                formula = cell.value if isinstance(cell.value, str) else f"={cell.value}"
                if not formula.startswith('='):
                    formula = f"={formula}"
                cells[key] = {"raw": formula, "type": "formula"}
            elif cell.value is not None and cell.value != '':
                if isinstance(cell.value, (int, float)):
                    cells[key] = {"raw": str(cell.value), "type": "number"}
                else:
                    cells[key] = {"raw": str(cell.value), "type": "text"}

            fmt = _parse_cell_format(cell)
            if fmt:
                formatting[key] = fmt

    return {
        "cells": cells,
        "formatting": formatting,
        "dimensions": {
            "rows": max(max_row + 1, 100),
            "cols": max(max_col + 1, 26),
        },
    }


def parse_xlsx_all_sheets(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse xlsx file and return all sheets."""
    workbook = load_workbook(filename=io.BytesIO(file_content), data_only=False)

    sheets = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        data = parse_worksheet(worksheet)
        sheets.append({
            "name": sheet_name,
            "data": data,
        })

    return sheets


def parse_csv(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse CSV file and convert to our cell format (single sheet)."""
    text = file_content.decode('utf-8')
    reader = csv.reader(io.StringIO(text))

    cells = {}
    max_row = 0
    max_col = 0

    for row_idx, row in enumerate(reader):
        for col_idx, value in enumerate(row):
            max_row = max(max_row, row_idx)
            max_col = max(max_col, col_idx)

            if value is None or value == '':
                continue

            key = f"{row_idx},{col_idx}"

            if value.startswith('='):
                cells[key] = {"raw": value, "type": "formula"}
            else:
                try:
                    float(value)
                    cells[key] = {"raw": value, "type": "number"}
                except ValueError:
                    cells[key] = {"raw": value, "type": "text"}

    return [{
        "name": "Sheet 1",
        "data": {
            "cells": cells,
            "dimensions": {
                "rows": max(max_row + 1, 100),
                "cols": max(max_col + 1, 26),
            },
        },
    }]


async def apply_template_to_workspace(template_name: str, workspace_id: str, user_id: str) -> dict:
    """Apply a template's sheets to an existing workspace (appending them)."""
    if not workspace_id:
        return {"error": "No workspace_id provided"}
    if not template_name:
        return {"error": "No template_name provided"}

    # Find template by name
    template = await execute_query_one(
        """SELECT id, name, preview_data FROM templates
           WHERE name = $1 AND (user_id IS NULL OR user_id = $2)""",
        template_name, user_id
    )
    if not template:
        return {"error": f"Template '{template_name}' not found"}

    # Verify user owns the workspace
    workspace = await execute_query_one(
        "SELECT id FROM workspaces WHERE id = $1 AND user_id = $2",
        workspace_id, user_id
    )
    if not workspace:
        return {"error": "Workspace not found"}

    # Fetch template sheets
    sheet_rows = await execute_query(
        """SELECT name, data, sort_order FROM template_sheets
           WHERE template_id = $1 ORDER BY sort_order ASC""",
        template["id"]
    )
    if not sheet_rows:
        return {"error": "Template has no sheets"}

    # Get existing sheet names in workspace for deduplication
    existing_sheets = await execute_query(
        "SELECT name FROM sheets WHERE workspace_id = $1",
        workspace_id
    )
    existing_names = {s["name"] for s in existing_sheets}

    # Insert template sheets
    created_sheets = []
    for sr in sheet_rows:
        data = sr["data"]
        if isinstance(data, str):
            data = json.loads(data)

        # Auto-rename if name conflicts
        name = sr["name"]
        if name in existing_names:
            counter = 2
            while f"{name} ({counter})" in existing_names:
                counter += 1
            name = f"{name} ({counter})"
        existing_names.add(name)

        sheet_id = secrets.token_urlsafe(12)
        await execute_command(
            """INSERT INTO sheets (id, workspace_id, user_id, name, data, updated_at)
               VALUES ($1, $2, $3, $4, $5::jsonb, NOW())""",
            sheet_id, workspace_id, user_id, name, data
        )
        created_sheets.append({"id": sheet_id, "name": name})

    # Copy template preview_data to workspace
    if template.get("preview_data"):
        await execute_command(
            "UPDATE workspaces SET preview_data = $1::jsonb, updated_at = NOW() WHERE id = $2 AND user_id = $3",
            template["preview_data"], workspace_id, user_id
        )

    return {
        "status": "applied",
        "template_name": template["name"],
        "sheets_added": [s["name"] for s in created_sheets],
    }


class CreateTemplateRequest(BaseModel):
    name: str
    sheets: List[Dict[str, Any]]  # [{name: str, data: {cells, dimensions}}, ...]
    thumbnail_url: Optional[str] = None


@router.get("/templates")
async def get_templates(user = Depends(get_workos_user)):
    """Get all templates (defaults + user's own) with sheet count."""
    user_id = user["id"]

    rows = await execute_query(
        """SELECT t.id, t.name, t.thumbnail_url, t.preview_data, t.user_id, t.created_at,
                  COUNT(ts.id) as sheet_count
           FROM templates t
           LEFT JOIN template_sheets ts ON ts.template_id = t.id
           WHERE t.user_id IS NULL OR t.user_id = $1
           GROUP BY t.id
           ORDER BY t.user_id NULLS FIRST, t.created_at DESC""",
        user_id
    )

    return [
        {
            "id": row["id"],
            "name": row["name"],
            "thumbnail_url": row["thumbnail_url"],
            "preview_data": row["preview_data"],
            "is_default": row["user_id"] is None,
            "sheet_count": row["sheet_count"],
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
        }
        for row in rows
    ]


@router.get("/templates/{template_id}")
async def get_template(
    template_id: int,
    user = Depends(get_workos_user)
):
    """Get a single template with its sheets."""
    user_id = user["id"]

    row = await execute_query_one(
        """SELECT id, name, thumbnail_url, user_id, created_at
           FROM templates
           WHERE id = $1 AND (user_id IS NULL OR user_id = $2)""",
        template_id, user_id
    )

    if not row:
        raise HTTPException(status_code=404, detail="Template not found")

    # Get sheets
    sheet_rows = await execute_query(
        """SELECT id, name, data, sort_order
           FROM template_sheets
           WHERE template_id = $1
           ORDER BY sort_order ASC""",
        template_id
    )

    sheets = []
    for sr in sheet_rows:
        data = sr["data"]
        if isinstance(data, str):
            data = json.loads(data)
        sheets.append({
            "id": sr["id"],
            "name": sr["name"],
            "data": data,
            "sort_order": sr["sort_order"],
        })

    return {
        "id": row["id"],
        "name": row["name"],
        "thumbnail_url": row["thumbnail_url"],
        "sheets": sheets,
        "is_default": row["user_id"] is None,
        "created_at": row["created_at"].isoformat() if row["created_at"] else None,
    }


@router.post("/templates/{template_id}/use")
async def use_template(
    template_id: int,
    user = Depends(get_workos_user)
):
    """Create a new workspace from a template with all its sheets."""
    user_id = user["id"]

    # Fetch template
    template = await execute_query_one(
        """SELECT id, name, preview_data
           FROM templates
           WHERE id = $1 AND (user_id IS NULL OR user_id = $2)""",
        template_id, user_id
    )

    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    # Fetch template sheets
    sheet_rows = await execute_query(
        """SELECT name, data, sort_order
           FROM template_sheets
           WHERE template_id = $1
           ORDER BY sort_order ASC""",
        template_id
    )

    if not sheet_rows:
        raise HTTPException(status_code=400, detail="Template has no sheets")

    # Create workspace
    workspace_id = secrets.token_urlsafe(12)
    await execute_command(
        """INSERT INTO workspaces (id, user_id, name, preview_data)
           VALUES ($1, $2, $3, $4::jsonb)""",
        workspace_id, user_id, template["name"], template.get("preview_data")
    )

    # Create sheets
    created_sheets = []
    for sr in sheet_rows:
        sheet_id = secrets.token_urlsafe(12)
        data = sr["data"]
        if isinstance(data, str):
            data = json.loads(data)

        await execute_command(
            """INSERT INTO sheets (id, workspace_id, user_id, name, data, updated_at)
               VALUES ($1, $2, $3, $4, $5::jsonb, NOW())""",
            sheet_id, workspace_id, user_id, sr["name"], data
        )
        created_sheets.append({"id": sheet_id, "name": sr["name"]})

    return {
        "workspace_id": workspace_id,
        "sheets": created_sheets,
    }


@router.post("/templates/upload")
async def upload_template(
    file: UploadFile = File(...),
    user = Depends(get_workos_user)
):
    """Upload and parse an xlsx or csv file as a template."""
    user_id = user["id"]

    if not file.filename:
        raise HTTPException(status_code=400, detail="Filename required")

    filename_lower = file.filename.lower()
    if not (filename_lower.endswith('.xlsx') or filename_lower.endswith('.csv')):
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are supported")

    content = await file.read()

    if len(content) > MAX_TEMPLATE_SIZE_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Max size is 1MB, got {len(content) / 1024 / 1024:.2f}MB"
        )

    try:
        if filename_lower.endswith('.xlsx'):
            sheets = parse_xlsx_all_sheets(content)
        else:
            sheets = parse_csv(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    # Validate total size
    total_size = sum(len(json.dumps(s["data"]).encode('utf-8')) for s in sheets)
    if total_size > MAX_TEMPLATE_SIZE_BYTES * 5:  # Allow up to 5MB for multi-sheet
        raise HTTPException(
            status_code=413,
            detail=f"Parsed data too large. Max size is 5MB, got {total_size / 1024 / 1024:.2f}MB"
        )

    name = file.filename.rsplit('.', 1)[0] if file.filename else "Untitled"

    # Extract preview_data from first sheet
    preview_data = None
    if sheets and sheets[0].get("data", {}).get("cells"):
        first_data = sheets[0]["data"]
        preview_data = extract_preview_data(first_data["cells"], first_data.get("formatting"))

    # Create template
    template_row = await execute_query_one(
        """INSERT INTO templates (name, user_id, preview_data)
           VALUES ($1, $2, $3::jsonb)
           RETURNING id, created_at""",
        name, user_id, preview_data
    )

    # Insert sheets
    for idx, sheet in enumerate(sheets):
        await execute_command(
            """INSERT INTO template_sheets (template_id, name, data, sort_order)
               VALUES ($1, $2, $3::jsonb, $4)""",
            template_row["id"], sheet["name"], sheet["data"], idx
        )

    return {
        "status": "created",
        "id": template_row["id"],
        "name": name,
        "sheet_count": len(sheets),
        "created_at": template_row["created_at"].isoformat() if template_row["created_at"] else None,
    }


@router.post("/templates")
async def create_template(
    body: CreateTemplateRequest,
    user = Depends(get_workos_user)
):
    """Create a new user template with sheets."""
    user_id = user["id"]

    if not body.sheets:
        raise HTTPException(status_code=400, detail="At least one sheet required")

    # Validate size
    total_size = sum(len(json.dumps(s.get("data", {})).encode('utf-8')) for s in body.sheets)
    if total_size > MAX_TEMPLATE_SIZE_BYTES * 5:
        raise HTTPException(
            status_code=413,
            detail=f"Template data too large. Max size is 5MB, got {total_size / 1024 / 1024:.2f}MB"
        )

    # Extract preview_data from first sheet
    preview_data = None
    if body.sheets and body.sheets[0].get("data", {}).get("cells"):
        first_data = body.sheets[0]["data"]
        preview_data = extract_preview_data(first_data["cells"], first_data.get("formatting"))

    # Create template
    template_row = await execute_query_one(
        """INSERT INTO templates (name, thumbnail_url, user_id, preview_data)
           VALUES ($1, $2, $3, $4::jsonb)
           RETURNING id, created_at""",
        body.name, body.thumbnail_url, user_id, preview_data
    )

    # Insert sheets
    for idx, sheet in enumerate(body.sheets):
        await execute_command(
            """INSERT INTO template_sheets (template_id, name, data, sort_order)
               VALUES ($1, $2, $3::jsonb, $4)""",
            template_row["id"], sheet.get("name", f"Sheet {idx + 1}"), sheet.get("data", {}), idx
        )

    return {
        "status": "created",
        "id": template_row["id"],
        "name": body.name,
        "sheet_count": len(body.sheets),
        "created_at": template_row["created_at"].isoformat() if template_row["created_at"] else None,
    }


@router.delete("/templates/{template_id}")
async def delete_template(
    template_id: int,
    user = Depends(get_workos_user)
):
    """Delete a user template (cannot delete default templates)."""
    user_id = user["id"]

    existing = await execute_query_one(
        "SELECT id, user_id FROM templates WHERE id = $1",
        template_id
    )

    if not existing:
        raise HTTPException(status_code=404, detail="Template not found")

    if existing["user_id"] is None:
        raise HTTPException(status_code=403, detail="Cannot delete default templates")

    if existing["user_id"] != user_id:
        raise HTTPException(status_code=403, detail="Not your template")

    # CASCADE will delete template_sheets
    await execute_command(
        "DELETE FROM templates WHERE id = $1",
        template_id
    )

    return {"status": "deleted", "id": template_id}
